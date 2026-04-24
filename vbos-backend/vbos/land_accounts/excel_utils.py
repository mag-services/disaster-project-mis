"""
Excel import/export for land accounts (opening and closing only).
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .constants import CATEGORIES, PROVINCES
from .utils import build_provinces_from_opening_closing


def create_template_workbook() -> Workbook:
    """Create a downloadable Excel template with Province, Category, Opening, Closing columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Land Accounts"
    # Headers
    ws.append(["Province", "Category", "Opening (km²)", "Closing (km²)"])
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = Font(bold=True)
    # Data rows: all province × category combinations
    for prov in PROVINCES:
        for cat in CATEGORIES:
            ws.append([prov, cat, 0, 0])
    # Instructions sheet (first so user sees it)
    inst = wb.create_sheet("Instructions", 0)
    inst.append(["Land Accounts Import Template"])
    inst.cell(row=1, column=1).font = Font(bold=True, size=14)
    inst.append([])
    inst.append(["Fill in Opening and Closing (km²) for each Province and Category."])
    inst.append(["The change matrix will be computed automatically on import."])
    inst.append([])
    inst.append(["Provinces:", ", ".join(PROVINCES)])
    inst.append(["Categories:", ", ".join(CATEGORIES)])
    inst.append([])
    inst.append(["Do not change the Province and Category values in the Land Accounts sheet."])
    wb.active = 1  # Land Accounts sheet is index 1
    return wb


def get_template_bytes() -> bytes:
    """Return template Excel file as bytes."""
    wb = create_template_workbook()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_import_file(file) -> dict:
    """
    Parse uploaded Excel file. Expected columns: Province, Category, Opening (km²), Closing (km²).
    Returns provinces dict for build_provinces_from_opening_closing.
    """
    wb = load_workbook(filename=file, read_only=True, data_only=True)
    # Find sheet with Province/Category headers
    ws = None
    for sheet in wb.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row and "province" in str(first_row[0] or "").lower():
            ws = sheet
            break
    if ws is None:
        ws = wb.active
    provinces_input = {p: {c: {"opening": 0, "closing": 0} for c in CATEGORIES} for p in PROVINCES}
    header = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        row = [str(c).strip() if c is not None else "" for c in row]
        if not any(row):
            continue
        if header is None:
            if "province" in (row[0] or "").lower() and "category" in (row[1] or "").lower():
                header = row
                continue
            continue
        prov = (row[0] or "").strip()
        cat = (row[1] or "").strip()
        if prov not in PROVINCES or cat not in CATEGORIES:
            continue
        try:
            opening = float(row[2]) if row[2] is not None and str(row[2]).strip() != "" else 0
            closing = float(row[3]) if row[3] is not None and str(row[3]).strip() != "" else 0
        except (ValueError, TypeError):
            opening = closing = 0
        provinces_input[prov][cat] = {"opening": opening, "closing": closing}
    wb.close()
    return provinces_input
