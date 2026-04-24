"""
Excel import/export for coastal changes (province, year, shoreline change).
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .constants import DEFAULT_YEARS, PROVINCES


def create_template_workbook() -> Workbook:
    """Create a downloadable Excel template with Province, Year, Shoreline change (m) columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Coastal Changes"
    # Headers
    ws.append(["Province", "Year", "Shoreline change (m)"])
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = Font(bold=True)
    # Data rows: all province × year combinations
    for prov in PROVINCES:
        for yr in DEFAULT_YEARS:
            ws.append([prov, yr, 0])
    # Instructions sheet (first so user sees it)
    inst = wb.create_sheet("Instructions", 0)
    inst.append(["Coastal Changes Import Template"])
    inst.cell(row=1, column=1).font = Font(bold=True, size=14)
    inst.append([])
    inst.append(["Fill in Shoreline change (m) for each Province and Year."])
    inst.append(["Positive = accretion, Negative = erosion."])
    inst.append([])
    inst.append(["Provinces:", ", ".join(PROVINCES)])
    inst.append(["Years:", ", ".join(DEFAULT_YEARS)])
    inst.append([])
    inst.append(["Do not change the Province and Year values in the Coastal Changes sheet."])
    wb.active = 1  # Coastal Changes sheet is index 1
    return wb


def get_template_bytes() -> bytes:
    """Return template Excel file as bytes."""
    wb = create_template_workbook()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_import_file(file) -> dict:
    """
    Parse uploaded Excel file. Expected columns: Province, Year, Shoreline change (m).
    Returns provinces dict: {province: {year: value}}.
    """
    wb = load_workbook(filename=file, read_only=True, data_only=True)
    ws = None
    for sheet in wb.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row and "province" in str(first_row[0] or "").lower():
            ws = sheet
            break
    if ws is None:
        ws = wb.active
    provinces = {p: {} for p in PROVINCES}
    header = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        row = [str(c).strip() if c is not None else "" for c in row]
        if not any(row):
            continue
        if header is None:
            if "province" in (row[0] or "").lower() and "year" in (row[1] or "").lower():
                header = row
                continue
            continue
        prov = (row[0] or "").strip()
        yr = str(row[1] or "").strip()
        if prov not in PROVINCES:
            continue
        try:
            value = float(row[2]) if row[2] is not None and str(row[2]).strip() != "" else 0
        except (ValueError, TypeError):
            value = 0
        provinces[prov][yr] = value
    wb.close()
    return {"provinces": provinces}
